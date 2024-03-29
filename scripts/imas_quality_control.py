#!/usr/bin/env python3

"""Extracts the photopeak from the energy per minimodule and saves it to a file.
Usage: imas_quality_control YAMLCONF INFILES ...

Arguments:
    YAMLCONF  File with all parameters to take into account in the scan.
    INFILE    Input file to be processed. Must be a compact binary file from PETsys.

Options:
    -h --help     Show this screen.    
"""


from collections import defaultdict
from itertools import chain
from multiprocessing import cpu_count, get_context
import os
import time
from typing import Callable
from docopt import docopt
from matplotlib import pyplot as plt
from matplotlib.colors import LogNorm
import numpy as np
import pandas as pd
import yaml
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from src.detector_features import calculate_centroid_sum, calculate_total_energy
from src.read_compact import read_binary_file
from src.filters import filter_min_ch
from src.mapping_generator import ChannelType, map_factory
from src.fits import fit_gaussian
from src.utils import get_max_en_channel, get_maxEnergy_sm_mM
from src.slab_nn import neural_net_pcalc, read_category_file


# Total number of eevents
EVT_COUNT_T = 0
# Events passing the filter
EVT_COUNT_F = 0


def increment_total():
    """
    Increments the total event count.

    This function is used to keep track of the total number of events processed.
    """
    global EVT_COUNT_T
    EVT_COUNT_T += 1


def increment_pf():
    """
    Increments the count of events that pass the filter.

    This function is used to keep track of the number of events that pass the filter.
    """
    global EVT_COUNT_F
    EVT_COUNT_F += 1


def extract_data_dict(
    read_det_evt: Callable,
    chtype_map: dict,
    sm_mM_map: dict,
    local_map: dict,
    min_ch: int,
    sum_rows_cols: bool,
    positions_pred: Callable,
    cat_map_XY: dict,
    cat_map_DOI: dict,
) -> tuple[dict, np.ndarray]:
    """
    This function processes the detector events and generates all the necessary dictionaries for the quality control.
    """
    bunch_size = 10000
    event_count = 0
    start_time = time.time()
    energy_ch_count = defaultdict(int)
    slab_dict_count = defaultdict(int)
    slab_dict_energy = defaultdict(list)
    sm_flood = defaultdict(list)
    infer_type = np.dtype([("slab_idx", np.int32), ("Esignals", np.float32, 8)])
    channel_energies = np.zeros(bunch_size, infer_type)
    event_energies = np.zeros(bunch_size, np.float32)
    # two columns with np.float32
    rtp_type = np.dtype([("x_rtp", np.float32), ("y_rtp", np.float32)])
    flood_rtp = np.zeros(bunch_size, rtp_type)
    cat_type = np.dtype([("Y", np.float32), ("DOI", np.float32)])
    cat_events = np.zeros(bunch_size, cat_type)
    chan_en_cnt = 0
    for event in read_det_evt:
        increment_total()
        # break to limit the number of events
        if EVT_COUNT_F > 2000000:
            break
        det1, det2 = event
        min_ch_filter1 = filter_min_ch(det1, min_ch, chtype_map, sum_rows_cols)
        min_ch_filter2 = filter_min_ch(det2, min_ch, chtype_map, sum_rows_cols)
        if EVT_COUNT_T % 100000 == 0:
            count_time = time.time()
            print(
                f"Events processed / passing filter: {EVT_COUNT_T} / {EVT_COUNT_F}\tTime taken: {round(count_time - start_time,1)} seconds"
            )
        if not (min_ch_filter1 and min_ch_filter2):
            continue
        # Get the mm with the maximum energy for each detector
        max_det1, energy_det1 = get_maxEnergy_sm_mM(det1, sm_mM_map, chtype_map)
        max_det2, energy_det2 = get_maxEnergy_sm_mM(det2, sm_mM_map, chtype_map)
        min_ch_maxdet1 = filter_min_ch(max_det1, min_ch, chtype_map, sum_rows_cols)
        min_ch_maxdet2 = filter_min_ch(max_det2, min_ch, chtype_map, sum_rows_cols)
        if not (min_ch_maxdet1 and min_ch_maxdet2):
            continue
        increment_pf()
        # Get the slab index (time channel) with the maximum energy in each detector
        slab_det1 = get_max_en_channel(max_det1, chtype_map, ChannelType.TIME)[2]
        slab_det2 = get_max_en_channel(max_det2, chtype_map, ChannelType.TIME)[2]
        slab_dict_count[slab_det1] += 1
        slab_dict_count[slab_det2] += 1
        slab_dict_energy[slab_det1].append(energy_det1)
        slab_dict_energy[slab_det2].append(energy_det2)

        x_det1, y_det1 = calculate_centroid_sum(
            max_det1, local_map, chtype_map, x_rtp=1, y_rtp=2
        )
        x_det2, y_det2 = calculate_centroid_sum(
            max_det2, local_map, chtype_map, x_rtp=1, y_rtp=2
        )
        flood_rtp[chan_en_cnt * 2]["x_rtp"] = x_det1
        flood_rtp[chan_en_cnt * 2]["y_rtp"] = y_det1
        flood_rtp[chan_en_cnt * 2 + 1]["x_rtp"] = x_det2
        flood_rtp[chan_en_cnt * 2 + 1]["y_rtp"] = y_det2

        # Fill the category per each of of the 10k events in the bunch
        try:
            cat_events[chan_en_cnt * 2]["Y"] = cat_map_XY[slab_det1]
        except KeyError:
            cat_events[chan_en_cnt * 2]["Y"] = 0

        try:
            cat_events[chan_en_cnt * 2]["DOI"] = cat_map_DOI[slab_det1]
        except KeyError:
            cat_events[chan_en_cnt * 2]["DOI"] = 0

        try:
            cat_events[chan_en_cnt * 2 + 1]["Y"] = cat_map_XY[slab_det2]
        except KeyError:
            cat_events[chan_en_cnt * 2 + 1]["Y"] = 0

        try:
            cat_events[chan_en_cnt * 2 + 1]["DOI"] = cat_map_DOI[slab_det2]
        except KeyError:
            cat_events[chan_en_cnt * 2 + 1]["DOI"] = 0

        en_ch_det1 = filter(lambda x: ChannelType.ENERGY in chtype_map[x[2]], max_det1)
        en_ch_det2 = filter(lambda x: ChannelType.ENERGY in chtype_map[x[2]], max_det2)

        # Fill the energy signals per each of the 10k events in the bunch
        for hit in en_ch_det1:
            pos = local_map[hit[2]][2]
            channel_energies[chan_en_cnt * 2]["slab_idx"] = slab_det1
            channel_energies[chan_en_cnt * 2]["Esignals"][pos] = hit[1]
            event_energies[chan_en_cnt * 2] = energy_det1
            energy_ch_count[hit[2]] += 1
        for hit in en_ch_det2:
            pos = local_map[hit[2]][2]
            channel_energies[chan_en_cnt * 2 + 1]["slab_idx"] = slab_det2
            channel_energies[chan_en_cnt * 2 + 1]["Esignals"][pos] = hit[1]
            event_energies[chan_en_cnt * 2 + 1] = energy_det2
            energy_ch_count[hit[2]] += 1

        # If the bunch is full, predict the positions
        if chan_en_cnt == bunch_size / 2 - 1:
            predicted_xy, doi = positions_pred(
                channel_energies["slab_idx"],
                channel_energies,
                cat_xy=cat_events["Y"],
                cat_doi=cat_events["DOI"],
            )
            for slab_id, xy, xyrtp, en in zip(
                channel_energies["slab_idx"], predicted_xy, flood_rtp, event_energies
            ):
                sm = sm_mM_map[slab_id][0]
                sm_flood[sm].append((xy[0], xy[1], xyrtp[0], xyrtp[1], slab_id, en))
            channel_energies = np.zeros(bunch_size, infer_type)
            flood_rtp = np.zeros(bunch_size, rtp_type)
            chan_en_cnt = 0
        chan_en_cnt += 1
    print("---------------------")
    end_time = time.time()
    print(len(slab_dict_count))
    print(f"Time taken: {end_time - start_time} seconds")
    print(f"Total events: {EVT_COUNT_T}")
    print(f"Events passing the filter: {EVT_COUNT_F}")
    return slab_dict_energy, sm_flood, energy_ch_count, EVT_COUNT_F


def extract_photopeak_slab(slab_dict_energy: dict) -> dict:
    """
    This function extracts the photopeak from the energy per minimodule.
    """
    slab_dict_photopeak = {}
    for key, value in tqdm(slab_dict_energy.items(), desc="Slab progress"):
        n, bins = np.histogram(value, bins=200, range=(0, 200))
        # n, bins, patches = plt.hist(value, bins=200, range=(0, 200))
        try:
            x, y, pars, _, _ = fit_gaussian(n, bins)
            mu, sigma = pars[1], pars[2]
        except RuntimeError:
            mu, sigma = 0, 0
            # plt.legend([f"Slab: {key}\nError fitting the Gaussian"])
            # plt.show()
        slab_dict_photopeak[key] = (mu, sigma)
        # plt.plot(x, y, "-r", label="fit")
        # plt.legend([f"Slab: {key}\nEnergy res: {round(2.35*sigma/mu*100,2)}%"])
        # plt.show()
    return slab_dict_photopeak


def slab_eval(
    slab_dict_photopeak: dict, chtype_map: dict, sm_mM_map: dict, eval_dir: str
) -> None:
    """
    This function evaluates the slab performance and if the channel is present or not in the system.
    """
    imas_lim_eres = 20
    time_channels = [
        ch for ch in sorted(chtype_map.keys()) if ChannelType.TIME in chtype_map[ch]
    ]
    non_working_tch = []
    out_of_range_tch = []
    for tch in sorted(time_channels):
        if tch in slab_dict_photopeak.keys():
            mu = slab_dict_photopeak[tch][0]
            sigma = slab_dict_photopeak[tch][1]
            if mu != 0:
                en_res = 2.35 * sigma / mu * 100
                if en_res > imas_lim_eres:
                    out_of_range_tch.append((tch, *sm_mM_map[tch], en_res))
            else:
                non_working_tch.append((tch, *sm_mM_map[tch]))

        else:
            non_working_tch.append((tch, *sm_mM_map[tch]))
    print(f"Number of missing time channels: {len(non_working_tch)}")
    print(f"Number of time channels out of range: {len(out_of_range_tch)}")

    # Create a DataFrame for all time channels
    df_all = pd.DataFrame(time_channels, columns=["tch"])
    df_all["SM"] = df_all["tch"].map(lambda x: sm_mM_map[x][0])
    df_all["mm"] = df_all["tch"].map(lambda x: sm_mM_map[x][1])
    df_all["Centroid"] = df_all["tch"].map(
        lambda x: slab_dict_photopeak.get(x, [0, 0])[0]
    )
    df_all["FWHM"] = df_all["tch"].map(
        lambda x: 2.35 * slab_dict_photopeak.get(x, [0, 0])[1]
    )
    df_all["Status"] = "Good"
    df_all.loc[df_all["tch"].isin([x[0] for x in non_working_tch]), "Status"] = (
        "Missing"
    )
    df_all.loc[df_all["tch"].isin([x[0] for x in out_of_range_tch]), "Status"] = (
        "Poor energy resolution"
    )

    # Create DataFrames for missing and out of range time channels
    df_missing = pd.DataFrame(non_working_tch, columns=["tch", "SM", "mm"])
    df_out_of_range = pd.DataFrame(out_of_range_tch, columns=["tch", "SM", "mm", "ER"])

    # Save the DataFrames to an Excel file
    with pd.ExcelWriter(f"{eval_dir}time_channels_QC.xlsx") as writer:
        df_all.to_excel(writer, sheet_name="All time channels", index=False)
        df_missing.to_excel(writer, sheet_name="Missing time channels", index=False)
        df_out_of_range.to_excel(
            writer, sheet_name="Poor energy resolution channels", index=False
        )

    # Load the workbook and get the sheet for all time channels
    wb = load_workbook(f"{eval_dir}time_channels_QC.xlsx")
    sheet = wb["All time channels"]

    # Define the fill colors
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    orange_fill = PatternFill(
        start_color="FFA500", end_color="FFA500", fill_type="solid"
    )

    # Apply the fill colors to the cells
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[3].value == "Missing":
            row[3].fill = red_fill
        elif row[3].value == "Poor energy resolution":
            row[3].fill = orange_fill

    # Save the workbook
    wb.save(f"{eval_dir}time_channels_QC.xlsx")


def energy_eval(
    energy_ch_dict: dict, chtype_map: dict, sm_mM_map: dict, eval_dir: str
) -> None:
    """
    This function evaluates the energy channels and if the channel is present or not in the system.
    """
    energy_channels = [
        ch for ch in sorted(chtype_map.keys()) if ChannelType.ENERGY in chtype_map[ch]
    ]
    non_working_ech = []
    for ech in sorted(energy_channels):
        if ech not in energy_ch_dict.keys():
            non_working_ech.append((ech, *sm_mM_map[ech]))
    print(f"Number of missing energy channels: {len(non_working_ech)}")

    # Create a DataFrame for all energy channels
    df_all = pd.DataFrame(energy_channels, columns=["ech"])
    df_all["SM"] = df_all["ech"].map(lambda x: sm_mM_map[x][0])
    df_all["mm"] = df_all["ech"].map(lambda x: sm_mM_map[x][1])
    df_all["Status"] = "Good"
    df_all.loc[df_all["ech"].isin([x[0] for x in non_working_ech]), "Status"] = (
        "Missing"
    )
    # Create DataFrames for missing and out of range energy channels
    df_missing = pd.DataFrame(non_working_ech, columns=["ech", "SM", "mm"])
    # Save the DataFrames to an Excel file
    with pd.ExcelWriter(f"{eval_dir}energy_channels_QC.xlsx") as writer:
        df_all.to_excel(writer, sheet_name="All energy channels", index=False)
        df_missing.to_excel(writer, sheet_name="Missing energy channels", index=False)

    # Load the workbook and get the sheet for all energy channels
    wb = load_workbook(f"{eval_dir}energy_channels_QC.xlsx")
    sheet = wb["All energy channels"]

    # Define the fill colors
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Apply the fill colors to the cells
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[3].value == "Missing":
            row[3].fill = red_fill

    # Save the workbook
    wb.save(f"{eval_dir}energy_channels_QC.xlsx")


def flood_map_eval(sm_flood: dict, slab_dict_photopeak: dict, eval_dir: str) -> None:
    """
    This function creates the flood map representation of the system.
    """
    fig1, axs1 = plt.subplots(5, 24, figsize=(100, 20), sharex=True, sharey=True)
    axs1 = axs1.ravel()  # Flatten the array of axes
    fig2, axs2 = plt.subplots(5, 24, figsize=(100, 20), sharex=True, sharey=True)
    axs2 = axs2.ravel()  # Flatten the array of axes

    for sm, flood in tqdm(sorted(sm_flood.items()), desc="SM progress"):
        i = sm
        x, y, xrtp, yrtp, slab_id, en = zip(*flood)
        x_filtered = []
        y_filtered = []
        xrtp_filtered = []
        yrtp_filtered = []
        for xit, yit, xrtp_it, yrtp_it, slab_idit, enit in zip(
            x, y, xrtp, yrtp, slab_id, en
        ):
            mu, sigma = slab_dict_photopeak[slab_idit]
            min_en = mu - 2 * sigma
            max_en = mu + 2 * sigma
            if enit < min_en or enit > max_en:
                continue
            x_filtered.append(xit)
            y_filtered.append(yit)
            xrtp_filtered.append(xrtp_it)
            yrtp_filtered.append(yrtp_it)
        x = x_filtered
        y = y_filtered
        xrtp = xrtp_filtered
        yrtp = yrtp_filtered

        # Create the individual plot for NN predicted positions
        plt.figure(figsize=(15, 15))
        plt.hist2d(
            x,
            y,
            bins=(200, 200),
            range=[[0, 108], [0, 108]],
            cmap="jet",
            norm=LogNorm(),
        )
        plt.colorbar(label="Counts")  # Add a colorbar
        plt.xlabel("X position (mm)", fontsize=14)
        plt.ylabel("Y position (mm)", fontsize=14)
        plt.xlim([0, 108])
        plt.ylim([0, 108])
        plt.title(f"Floodmap Representation of sm {sm} NN", fontsize=16)
        plt.savefig(f"{eval_dir}floodmapCAT_sm{sm}.png", dpi=300)
        plt.close()  # Close the individual plot

        # Create the individual plot for the rtp positions
        plt.figure(figsize=(15, 15))
        plt.hist2d(
            xrtp,
            yrtp,
            bins=(200, 200),
            range=[[0, 108], [0, 108]],
            cmap="jet",
            norm=LogNorm(),
        )
        plt.colorbar(label="Counts")  # Add a colorbar
        plt.xlabel("X position (mm)", fontsize=14)
        plt.ylabel("Y position (mm)", fontsize=14)
        plt.xlim([0, 108])
        plt.ylim([0, 108])
        plt.title(f"Floodmap Representation of sm {sm} RTP", fontsize=16)
        plt.savefig(f"{eval_dir}floodmap_smCAT{sm}_rtp.png", dpi=300)
        plt.close()  # Close the individual plot

        # Add the data to the combined plot with rtp
        axs1[i].hist2d(
            x,
            y,
            bins=(200, 200),
            range=[[0, 108], [0, 108]],
            cmap="jet",
            norm=LogNorm(),
        )
        axs1[i].set_xlim([0, 108])
        axs1[i].set_ylim([0, 108])
        axs1[i].set_title(f"sm {sm}")

        # Add the data to the combined plot with rtp
        axs2[i].hist2d(
            xrtp,
            yrtp,
            bins=(200, 200),
            range=[[0, 108], [0, 108]],
            cmap="jet",
            norm=LogNorm(),
        )
        axs2[i].set_xlim([0, 108])
        axs2[i].set_ylim([0, 108])
        axs2[i].set_title(f"sm {sm}")

    # Adjust the space between the subplots
    plt.figure(fig1.number)
    plt.subplots_adjust(wspace=0, hspace=0)
    plt.tight_layout()
    plt.savefig(f"{eval_dir}floodmapCAT_fullIMAS.png", dpi=300)

    plt.figure(fig2.number)
    plt.subplots_adjust(wspace=0, hspace=0)
    plt.tight_layout()
    plt.savefig(f"{eval_dir}floodmapCAT_fullIMAS_rtp.png", dpi=300)


def initialize_nn(local_map: dict) -> Callable:
    """
    This function initializes the neural network for the position prediction.
    """
    nn_yfile = "/scratch/imas_files_cal/IMASY.h5"
    nn_doifile = "/scratch/imas_files_cal/IMASDOI.h5"
    positions_pred = neural_net_pcalc("IMAS", nn_yfile, nn_doifile, local_map)
    return positions_pred


def process_file(
    binary_file_path: str,
    chtype_map: dict,
    sm_mM_map: dict,
    local_map: dict,
    min_ch: int,
    sum_rows_cols: bool,
    positions_pred: Callable,
    cat_map_XY: dict,
    cat_map_DOI: dict,
) -> dict:
    """
    This function processes the binary file and returns a dictionary of energy list values for each slab.
    """

    print(f"Processing file: {binary_file_path}")
    # Read the binary file
    reader = read_binary_file(binary_file_path)
    # Extract the data dictionary
    slab_dict_energy = extract_data_dict(
        reader,
        chtype_map,
        sm_mM_map,
        local_map,
        min_ch,
        sum_rows_cols,
        positions_pred,
        cat_map_XY,
        cat_map_DOI,
    )
    return slab_dict_energy


def main():
    # Read the YAML configuration file
    args = docopt(__doc__)

    # Read the binary file
    binary_file_paths = args["INFILES"]

    with open(args["YAMLCONF"], "r") as f:
        config = yaml.safe_load(f)

    # Read the mapping file
    map_file = config["map_file"]

    # Get the coordinates of the channels
    local_map, sm_mM_map, chtype_map, FEM_instance = map_factory(map_file)

    sum_rows_cols = FEM_instance.sum_rows_cols

    min_ch = int(config["min_ch"])
    print(f"Minimum number of channels: {min_ch}")

    en_min_ch = float(config["en_min_ch"])
    print(f"Minimum energy per channel: {en_min_ch}")

    cat_map_XY = read_category_file("/scratch/imas_files_cal/CategoryMapRTPY.bin")
    cat_map_DOI = read_category_file("/scratch/imas_files_cal/CategoryMapRTPDOI.bin")
    print(f"Category maps read.")

    positions_pred = initialize_nn(local_map)

    with get_context("spawn").Pool(processes=cpu_count()) as pool:
        args_list = [
            (
                binary_file_path,
                chtype_map,
                sm_mM_map,
                local_map,
                min_ch,
                sum_rows_cols,
                positions_pred,
                cat_map_XY,
                cat_map_DOI,
            )
            for binary_file_path in binary_file_paths
        ]
        results = pool.starmap(process_file, args_list)

    slab_dict_energy = defaultdict(list)
    sm_dict_flood = defaultdict(list)
    energy_ch_dict = defaultdict(int)
    total_number_events = 0
    print(f"Processing results. Adding to the final dictionary, please wait...")
    for result in tqdm(results):
        if isinstance(result, Exception):
            print(f"Exception in child process: {result}")
        else:
            slab_dict, sm_dict, energy_dict, evt_filtered = result
            total_number_events += evt_filtered
            for key, value in slab_dict.items():
                slab_dict_energy[key].extend(value)
            for key, value in sm_dict.items():
                sm_dict_flood[key].extend(value)
            for key, value in energy_dict.items():
                energy_ch_dict[key] += value

    print(f"Total number of events processed: {total_number_events}")

    eval_dir = "imas_QC/"
    if not os.path.exists(eval_dir):
        os.makedirs(eval_dir)

    print(f"Extracting photopeak from the energy per minimodule...")
    slab_dict_photopeak = extract_photopeak_slab(slab_dict_energy)
    print(f"Photopeak extraction finished.")
    print(f"Processing flood map...")
    flood_map_eval(sm_dict_flood, slab_dict_photopeak, eval_dir)
    print(f"Flood map processing finished.")
    print(f"Evaluating the slab performance...")
    slab_eval(slab_dict_photopeak, chtype_map, sm_mM_map, eval_dir)
    print(f"Slab performance evaluation finished.")
    print(f"Evaluating the energy channels...")
    energy_eval(energy_ch_dict, chtype_map, sm_mM_map, eval_dir)
    print(f"Energy channel evaluation finished.")
    print(f"Quality control finished.")


if __name__ == "__main__":
    main()
